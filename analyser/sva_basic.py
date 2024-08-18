from .sva_plots import SvaPlots
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
import shutil

class SvaBasics(SvaPlots):
    TABLE_SIZE = None
    NUM_PLAYERS_GRAPH = 5


    def __init__(self, data, excelPath) -> None:
        super().__init__()
        self.rawData = data
        #self.excel_delimiter = excel_delimiter
        self.seasonDf = None
        self.sheetData = {}
        self.excelPath = excelPath

        if not os.path.isfile(excelPath):
            shutil.copyfile('statistik_template.xlsx', self.excelPath)

    
    def common_data_preparation(sef, df:pd.DataFrame):
        """
            data: DataFrame
        """
        df['date'] = pd.to_datetime(df['date'])
        df['weekday'] = df['date'].dt.day_name(locale='de_DE.utf-8')
        df = df.sort_values('date').reset_index()
        return df.copy()

    def get_season_df(self, season):
        season_id = season['id']

        return self.df.loc[self.df['season.id'] == season_id].copy()
    
    def toExcelSheet(self, df:pd.DataFrame, sheetName:str, index=False):
        writer = pd.ExcelWriter(self.excelPath, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        
        df.to_excel(writer, sheet_name=sheetName , startrow=0, header=True, index=index)
        # Get the xlsxwriter workbook and worksheet objects.
        worksheet = writer.sheets[sheetName]

        tableName =sheetName.replace(' ', '_')

        if tableName in worksheet.tables:
            del worksheet.tables[tableName]

        # Get the dimensions of the dataframe.
        (max_row, max_col) = df.shape
        
        if index:
            max_col += 1
        
        # When A-Z
        if max_col <= 24:
            maxColString = chr(max_col+64)
        else:
            # When over Z
            maxColString = 'A' + chr(64 + max_col%26)

        ref = 'A1:{0}{1}'.format(maxColString, str(max_row +1))
        #print(ref)
        tab = Table(displayName=tableName, ref=ref)

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        # # Create a list of column headers, to use in add_table().

        # # Add the Excel table structure. Pandas will add the data.
        # worksheet.add_table(0, 0, max_row, max_col - 1, {"columns": column_settings})
        worksheet.add_table(tab)
        
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length
        
        # Remove former columns (in case the table got smaller)
        worksheet.delete_cols(max_col +1, (worksheet.max_column - max_col))
        # Remove former rows (in case the table got smaller)
        worksheet.delete_rows(max_row+2, worksheet.max_row - max_row)
        #writer.save()
        writer.close()